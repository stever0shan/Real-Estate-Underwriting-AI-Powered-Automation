from openpyxl import load_workbook

# These are guessed placements — can be tweaked after visual confirmation
FIELD_CELL_MAP = {
    "monthly_rent": "B15",     
    "NOI": "B8",
    "cap_rate": "B12",
    "total_units": "B5"
}

def fill_excel(template_path, output_path, field_data):
    wb = load_workbook(template_path)
    if "Project Summary" in wb.sheetnames:
        ws = wb["Project Summary"]
    else:
        raise Exception("Expected 'Project Summary' tab not found in template")

    for field, cell in FIELD_CELL_MAP.items():
        value = field_data.get(field, "NEEDS_REVIEW")
        ws[cell] = value

    wb.save(output_path)
