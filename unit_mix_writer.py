from openpyxl import load_workbook
import re

def is_valid_row(row):
    return any(cell and str(cell).strip() not in ["", "-", "None"] for cell in row)

def clean_value(val):
    if isinstance(val, str):
        val = val.replace("$", "").replace(",", "").strip()
        if val.endswith(".00"): 
            val = val[:-3]
    return val

def add_unit_mix_tab(excel_path, tables):
    wb = load_workbook(excel_path)
    ws = wb.create_sheet("Unit Mix") if "Unit Mix" not in wb.sheetnames else wb["Unit Mix"]

    ws.append(["Unit Type", "Unit Count", "Avg SF", "Monthly Rent", "Rent/SF"])

    for table in tables:
        df = table.df
        for row in df.itertuples(index=False):
            values = [str(cell).strip() for cell in row]
            if is_valid_row(values):
                cleaned = [clean_value(val) for val in values[:5]]
                ws.append(cleaned)

    wb.save(excel_path)
